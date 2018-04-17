
__author__ = 'Daniel Beganyi'

# Import config file, set working directory.
from invest_config import *
import os
if working_dir:
    os.chdir(working_dir)

# Import modules datetime and openpyxl.
import datetime
import openpyxl
from openpyxl.utils import coordinate_from_string
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.table import Table

print("Modules imported.\nWorking directory: ", os.getcwd())

# Load excel file.
excel_file = openpyxl.load_workbook(file_to_load)

# Set present date as present.
present = datetime.date.today()

# Create class Portfolio_data
class Portfolio_data():

    def __init__(self, sheet, tableName):
        self.sheet = sheet
        self.tableName = tableName

    # Insterts rows: above x-th row, inster y number of rows.
    def insert_rows(self, x, y):
        self.sheet.insert_rows(x, amount=y)

    # Puts empty values in column A, starting in A2
    def clear_date_column(self):
        for row in self.sheet.iter_rows(min_row=2, min_col=1, max_col=1, max_row=self.sheet.max_row):
            for cell in row:
                cell.value = ''

    # Clears rest of sheet (columns B-- )
    def clear_numbers(self):
        for row in self.sheet.iter_rows(min_row=2, min_col=2, max_col=self.last_column_of_table('number'), max_row=self.sheet.max_row):
            for cell in row:
                cell.value = ''

    # Returns index of last column of table, in 'letter' or 'number'  format.
    def last_column_of_table(self, format):
        for i, table in enumerate(self.sheet._tables):
            if table.displayName == self.tableName:
                old_table_reference = table.ref
        last_cell_of_table = old_table_reference[old_table_reference.find(':') + 1:]
        last_column_letter = coordinate_from_string(last_cell_of_table)[0]
        if format == 'letter':
            return last_column_letter
        elif format == 'number':
            return column_index_from_string(last_column_letter)

    # Returns last row number in column A based on last valid date value.
    def last_row_in_date_column(self):
        n = 1
        for row in self.sheet.iter_rows(min_row=2, min_col=1, max_col=1, max_row=self.sheet.max_row):
            for cell in row:
                if type(cell.value) is datetime.date or type(cell.value) is datetime.datetime:
                    n += 1
        return n

    # Sets new end coordinates for a table, based on length of data in the worksheet.
    def modify_table_range(self):
        for i, table in enumerate(self.sheet._tables):
            if table.displayName == self.tableName:
                tableCnt = i
        new_table_reference = 'A1:' + self.last_column_of_table('letter') + str(self.last_row_in_date_column())
        newTable = Table(displayName=self.tableName, ref=new_table_reference)
        self.sheet._tables[tableCnt] = newTable

    # Fills column A with the dates between today and a past startdate. Reverse order (A2 is today).
    # Different startdate can be given as parameter, but it should be: latest date from the previous update + 1 day.
    # See below: 'setup_dates' function
    def fill_date_column(self, startdate):
        date_to_print = present
        for row in self.sheet.iter_rows(min_row=2, min_col=1, max_col=1, max_row=self.interval() + 1):
            for cell in row:
                cell.value = date_to_print
            if date_to_print == self.startdate():
                break
            else:
                date_to_print = date_to_print - datetime.timedelta(days=1)
        print("Date column filled.")

    # Sets type in date column 'datetime.date'.
    def set_date_type_for_date_column(self):
        for row in self.sheet.iter_rows(min_row=2, min_col=1, max_col=1, max_row=self.sheet.max_row):
            for cell in row:
                if cell.value == datetime.datetime:
                    cell.value = cell.value.date()

    # Fills each column under ticker with the online quote corresponding to the reference date in col A.
    # Empty cells (weekend) will be overwritten with the last available quote.
    def fill_quote_sheet(self):

        # Iterate through columns under Tickers
        for col in self.sheet.iter_cols(min_row=1, min_col=2, max_col=self.last_column_of_table('number'), max_row=1):
            colnum = (column_index_from_string(col[0].column))
            # Set ticker
            ticker = self.sheet.cell(column=colnum, row=1).value
            if ticker == '':
                print("Header cell is empty!")
                pass
            print(colnum - 1, "-", ticker)

            # Get API data
            quote_data = get_quote_json(ticker)
            quote_dates_list = list(quote_data.keys())

            # This loop will repeat 7 times, so (blank) cells for weekend days get written the last available data.
            for i in range(7):
                for row in self.sheet.iter_rows(min_row=2, min_col=colnum, max_col=colnum,
                                                    max_row=self.interval() + 1):
                    for cell in row:
                        # Get the reference date from col A.
                        reference_date = str(self.sheet.cell(column=1, row=cell.row).value)
                        # Write quote for date in cell, if available.
                        if reference_date in quote_dates_list:
                            if quote_data[reference_date]["4. close"] != "0.0000":
                                    cell.value = float(quote_data[reference_date]["4. close"])
                        # If cell is blank, try to get quote from the next cell
                        if cell.value == None:
                            cell.value = cell.offset(row=+1).value

    # Fills each row with a currency pair figure, corresponding to the reference date in col A.
    def fill_forex_sheet(self):

        # Iterates through USD base columns first (4-6).
        for row in self.sheet.iter_rows(min_row=2, min_col=4, max_col=6, max_row=self.interval() + 1):
            reference_date = str(self.sheet.cell(column=1, row=row[0].row).value)
            for cell in row:
                ticker = self.sheet.cell(column=column_index_from_string(cell.column), row=1).value
                cell.value = float(get_forex_json(reference_date)[ticker])

        # Iterates and fills HUF base columns 2-3 with formulas, to calculate a price from columns 4-6.
        for row in self.sheet.iter_rows(min_row=2, min_col=2, max_col=3, max_row=self.interval() + 1):
            for cell in row:
                if cell.column == 'B':
                    cell.value = '=INDIRECT("RC[2]",0)/INDIRECT("RC[3]",0)'
                elif cell.column == 'C':
                    cell.value = '=INDIRECT("RC[1]",0)/INDIRECT("RC[3]",0)'

    # Latest date: ideally the date of the last update. Value of cell A2, in default.
    def get_latest_previous_date(self):
        if type(self.sheet[first_cell_of_date_column].value) is datetime.datetime:
            return self.sheet[first_cell_of_date_column].value.date()
        elif type(self.sheet[first_cell_of_date_column].value) is datetime.date:
            return self.sheet[first_cell_of_date_column].value
        else:
            return False

    # Startdate = date from wich to fill date col and get corresponding online data. Latest date + 1 day.
    def startdate(self):
        return self.latest_previous_date + datetime.timedelta(days=1)


    # Interval = number of days to write and get data for.
    def interval(self):
        return (present - self.startdate() + datetime.timedelta(days=1)).days


# Fills cells in column B-- ,with online data.
def fill_sheet(ws):
    if ws == quote:
        ws.fill_quote_sheet()
    elif ws == forex:
        ws.fill_forex_sheet()
    # New sheet & function can be added here

# Saves file under previous name, or newly given filename.
def save_to_xlsx_file():
    filename = file_to_load

    answer = input("OK to overwrite original file: {filename}, or save as another name? ( y / n ) ")
    if answer not in ("y", " ", ""):
        filename = input("Input filename (with .xlsx) to save: ")

    finished = False
    while finished == False:
        try:
            excel_file.save(filename)
            print(f"\nExcel saved to {os.getcwd()} as {filename}")
            finished = True
        except PermissionError:
            answer = input("\nCould not save data. Close file! \n Try again? ( y / n ) ")
            if answer == "n":
                print("\nData NOT saved. Quitting.")
                finished = True

                
'''
EXECUTIVE PART STARTS HERE.
'''

# Create class instances.
quote = Portfolio_data(excel_file[quote_sheet_name], quote_table_name)
forex = Portfolio_data(excel_file[forex_sheet_name], forex_table_name)

update_happened = False

# Check if latest date is earlier than today, decide to update or pass.
for ws in [quote, forex]:

    ws.latest_previous_date = ws.get_latest_previous_date()

    if ws.latest_previous_date:
        
        if ws.latest_previous_date < present:
            
            print(f"\nLatest date in {ws.tableName} is {ws.latest_previous_date}.\nUpdating.")
            ws.insert_rows(2, ws.interval())
            ws.fill_date_column(ws.startdate())
            ws.set_date_type_for_date_column()
            ws.modify_table_range()
            
            try:
                fill_sheet(ws)
            except (KeyError, TypeError):
                print("\nKeyError in getting data from API."
                      "\nMissing or invalid access key?")
                print(f"{ws.tableName} skipped.")
                continue
            except requests.exceptions.HTTPError as e:
                print(e)
                print(f"{ws.tableName} skipped.")
                continue
                
            update_happened = True
            print(f"{ws.tableName} filled.")

        elif ws.latest_previous_date == present:
            print(f"\nLatest date in {ws.tableName} is present date.\nUpdate skipped.")

        elif ws.latest_previous_date > present:
            ws.latest_previous_date = False

    if ws.latest_previous_date is False:
        
        answer = False
        while answer not in ("c", "p"):
            answer = input(f"\nValue in cell {first_cell_of_date_column} in {ws.tableName} is '{ws.get_latest_previous_date()}'. "
                           f"  (invalid: empty, wrong type, or a future date)."
                           f"\nDo you want to: "
                           f"\n- Clear all data in {ws.tableName} and get new data, starting from a date of your input ( c ) "
                           f"\n - Pass ( p ) ?  ")
        
        if answer == "p":
            print("Choice: Pass")
            continue
            
        elif answer == "c":
            
            date_input = input("Input starting date (YYYY-MM-DD) :  ")
            try:
                ws.latest_previous_date = datetime.datetime.strptime(date_input, "%Y-%m-%d").date() - datetime.timedelta(
                days=1)
            except ValueError:
                print("Invalid input. Update skipped.")
                continue
            if ws.latest_previous_date >= present:
                print("Date entered is a future date. Update skipped.")
                continue
                
            print("\nClearing sheet and updating.")
            ws.clear_date_column()
            ws.clear_numbers()
            ws.fill_date_column(ws.startdate())
            ws.set_date_type_for_date_column()
            ws.modify_table_range()
            
            try:
                fill_sheet(ws)
            except (KeyError, TypeError):
                print("\nERROR in getting data from API."
                      "\nMissing or invalid access key?")
                print(f"{ws.tableName} skipped.")
                continue
            except requests.exceptions.HTTPError as e:
                print(e)
                print(f"{ws.tableName} skipped.")
                continue
                
            update_happened = True
            print(f"{ws.tableName} filled.")


if update_happened == True:
    save_to_xlsx_file()
else:
    print("\nNo update in data. File not saved. Quitting.")
